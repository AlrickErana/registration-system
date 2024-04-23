from tkinter import *
import datetime
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl, xlrd
from openpyxl import Workbook, worksheet
import pathlib


background = "#06283D"
framebg = "#EDEDED"
framefg = "#06283D"

root = Tk()
root.title("Student Registration system")
root.geometry("1250x700+210+100")
root.config(bg=background)

#15:32

# Check if the file exists
file_path = 'C:/Users/colis/Documents/tetris/registration/Student_data.xlsx'
file = pathlib.Path(file_path)
if file.exists():
    pass
else:
    if not file.exists():
        # Create a new workbook if the file doesn't exist
        file = Workbook()
        sheet = file.active
        sheet['A1'] = "Registration No."
        sheet['B1'] = "Name"
        sheet['C1'] = "Class"
        sheet['D1'] = "Gender"
        sheet['E1'] = "DOB"
        sheet['F1'] = "Date of Registration"
        sheet['G1'] = "Religion"
        sheet['H1'] = "Skill"
        sheet['I1'] = "Father Name"
        sheet['J1'] = "Mother Name"
        sheet['K1'] = "Father's Occupation"
        sheet['L1'] = "Mother's Occupation"
              
        # Save the workbook
        file.save(file_path)

def Exit():
    root.destroy()

def showimage():
    global filename
    global img
    filename = filedialog.askopenfilename(initialdir=os.getcwd(),
                                          title = "Select image file", filetype=(
                                                                                 ("PNG File", "*.png"),
                                                                                 ("JPG File", '*.jpg')
                                                                                 ))
    img = (Image.open(filename))
    resized_image = img.resize((190, 190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image=photo2


#Registration N0.
def registration_no():
    file = openpyxl.load_workbook(file_path)
    sheet = file.active
    row = sheet.max_row

    max_row_value = sheet.cell(row=row, column=1).value
    try:
        registration.set(max_row_value + 1)
    except:
        registration.set("1")


#clear
def clear():
    name.set('')
    dob.set('')
    religion.set('')
    skill.set('')
    F_Name.set('')
    m_Name.set('')
    father_occupation.set('')
    mother_occupation.set('')
    Class.set("Select Class")

    registration_no()

    savebutton.config(state='normal')

    img1 = PhotoImage(file='C:/Users/colis/Documents/tetris/registration/images/profile4.png')
    lbl.config(image=img1)
    lbl.image = img1

    
def Search():
    text = search.get()

    clear()
    savebutton.config(state='disable')

    file = openpyxl.load_workbook(file_path)
    sheet = file.active

    x1, x2, x3, x4, x5, x6, x7, x8, x9, x10, x11, x12 = None, None, None, None, None, None, None, None, None, None, None, None
#iter_rows
    if len(text) == 0:
        messagebox.showerror("Invalid",'Please Enter A Valid Registration Number')
    else:
        for row in sheet.iter_rows(values_only=True):
            if row[0] == int(text):
                x1, x2, x3, x4, x5, x6, x7, x8, x9, x10, x11, x12 = row
                break
            

        if x1 is None:
            messagebox.showerror("Invalid", "Invalid Registration Number!!!")
            return
        

        registration.set(x1)
        name.set(x2)
        Class.set(x3)
        if x4 == 'Female':
            r2.select()
        else:
            r1.select()
        dob.set(x5)
        Date.set(x6)
        religion.set(x7)
        skill.set(x8)
        F_Name.set(x9)
        m_Name.set(x10)
        father_occupation.set(x11)
        mother_occupation.set(x12)


        img = Image.open("C:/Users/colis/Documents/tetris/registration/student_profiles/" + str(x1) + '.png' or '.jpg')

        # img = Image.open(str(x1))

        resixed_image = img.resize((190, 190))
        photo2 = ImageTk.PhotoImage(resixed_image)
        lbl.config(image=photo2)
        lbl.image = photo2

#################################

def update():
    name = StringVar()
    r1 = registration.get()
    n1 = name.get()
    c1 = Class.get()
    selection()
    g1 = gender
    d2 = dob.get()
    d1 = Date.get()
    re1 = religion.get()
    s1 = skill.get()
    father_name = F_Name.get()
    mother_name = m_Name.get()
    f1 = father_occupation.get()
    m1 = mother_occupation.get()
    
    file = openpyxl.load_workbook(file_path)
    sheet = file.active

    for row in sheet.rows:
        if row[0].value == r1:
            name = row[0]
            reg_no_position = str(name)[14:-1]
            reg_numbers = str(name)[15:-1]

    # sheet.cell(column=1, row=int(reg_numbers), value=r1)
    sheet.cell(column=2, row=int(reg_numbers), value=n1)
    sheet.cell(column=3, row=int(reg_numbers), value=c1)
    sheet.cell(column=4, row=int(reg_numbers), value=g1)
    sheet.cell(column=5, row=int(reg_numbers), value=d2)
    sheet.cell(column=6, row=int(reg_numbers), value=d1)
    sheet.cell(column=7, row=int(reg_numbers), value=re1)
    sheet.cell(column=8, row=int(reg_numbers), value=s1)
    sheet.cell(column=9, row=int(reg_numbers), value=father_name)
    sheet.cell(column=10, row=int(reg_numbers), value=mother_name)
    sheet.cell(column=11, row=int(reg_numbers), value=f1)
    sheet.cell(column=12, row=int(reg_numbers), value=m1)

    file.save('C:/Users/colis/Documents/tetris/registration/Student_data.xlsx')

    try:
        img.save("C:/Users/colis/Documents/tetris/registration/student_profiles/" + str(r1) + '.png' or '.jpg')
    except:
        pass
    messagebox.showinfo("Update", "Update Successfully!!!")

    clear()

#saving
def Save():
    r1 = registration.get()
    n1 = name.get()
    c1 = Class.get()
    try:
        g1 = gender
    except:
        messagebox.showerror("Error", "Select Gender")

    d2 = dob.get()
    d1 = Date.get()
    re1 = religion.get()
    s1 = skill.get()
    father_name = F_Name.get()
    mother_name = m_Name.get()
    f1 = father_occupation.get()
    m1 = mother_occupation.get()

    if n1 == "" or c1 == "Select Class" or d2 == "" or re1 == "" or father_name == "" or mother_name == "" or m1 == "":
        messagebox.showerror("Error", "Few Data is missing")
    else:
        file = openpyxl.load_workbook(file_path)
        sheet = file.active
        sheet.cell(column=1, row=sheet.max_row+1, value=r1)
        sheet.cell(column=2, row=sheet.max_row, value=n1)
        sheet.cell(column=3, row=sheet.max_row, value=c1)
        sheet.cell(column=4, row=sheet.max_row, value=g1)
        sheet.cell(column=5, row=sheet.max_row, value=d2)
        sheet.cell(column=6, row=sheet.max_row, value=d1)
        sheet.cell(column=7, row=sheet.max_row, value=re1)
        sheet.cell(column=8, row=sheet.max_row, value=s1)
        sheet.cell(column=9, row=sheet.max_row, value=father_name)
        sheet.cell(column=10, row=sheet.max_row, value=mother_name)
        sheet.cell(column=11, row=sheet.max_row, value=f1)
        sheet.cell(column=12, row=sheet.max_row, value=m1)
        file.save(file_path)

        try:
            img.save("C:/Users/colis/Documents/tetris/registration/student_profiles/" + str(r1) + ".png")
            
        except:
            messagebox.showinfo("Info", "Profile Picture is not available")
        messagebox.showinfo("Info", "Successfully Data Entered")
        clear()
        registration_no()


def selection():
    global gender
    value = radio.get()
    if value == 1:
        gender = "Male"
    else:
        gender = "Female"

Label(root, text = "Email: example@gmail.com", width=10, height=3, bg="#f0687c", anchor='e').pack(side=TOP, fill=X)
Label(root, text = "STUDENT REGISTRATION", width=10, height=2, bg="#c36464", fg="#fff", font=("arial 20 bold")).pack(side=TOP, fill=X)

#search box to update
search = StringVar()
Entry(root, textvariable=search, width=15, bd=2, font="arial 20").place(x=820, y=70)
imageicon3 = PhotoImage(file="C:/Users/colis/Documents/tetris/registration/images/search2.png")
srch = Button(root, text='Search', compound=LEFT, image=imageicon3, width=123, bg='#68ddfa', font="arial 13 bold", command=Search)
srch.place(x=1060, y=70)

imageicon4 = PhotoImage(file='C:/Users/colis/Documents/tetris/registration/images/layer.png')
update_button = Button(root, image=imageicon4, height=30, width=70, bg='#c36464', command=update)
update_button.place(x=110, y=70)

Label(root, text="Registration No:", font='arial 13', fg=framebg, bg=background).place(x=30, y=150)
Label(root, text="Date:", font='arial 13', fg=framebg, bg=background).place(x=500, y=150)

registration = IntVar()
Date = StringVar()

reg_entry = Entry(root, textvariable=registration, width=15, font="arial 10")
reg_entry.place(x=160, y=150)
registration_no()

today = date.today()
d1 = today.strftime("%d/%m/%Y")
date_entry = Entry(root, textvariable=Date, width=15, font="arial 10")
date_entry.place(x=550, y=150)

Date.set(d1)

#student detail
obj = LabelFrame(root, text="Student's Details", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=250, relief=GROOVE)
obj.place(x=30, y=200)

Label(obj, text="Full Name:", font="Arial 13", bg=framebg, fg=framefg).place(x=30, y=50)
Label(obj, text="Date of Birth:", font="Arial 13", bg=framebg, fg=framefg).place(x=30, y=100)
Label(obj, text="Gender:", font="Arial 13", bg=framebg, fg=framefg).place(x=30, y=150)

Label(obj, text="Class", font="Arial 13", bg=framebg, fg=framefg).place(x=500, y=50)
Label(obj, text="Religion:", font="Arial 13", bg=framebg, fg=framefg).place(x=500, y=100)
Label(obj, text="Skills:", font="Arial 13", bg=framebg, fg=framefg).place(x=500, y=150)

name = StringVar()
name_entry = Entry(obj, textvariable=name, width=20, font="arial 10")
name_entry.place(x=160, y=50)

dob = StringVar()
dob_entry = Entry(obj, textvariable=dob, width=20, font="arial 10")
dob_entry.place(x=160, y=100)

radio = IntVar()
r1 = Radiobutton(obj, text="Male", variable=radio, value=1, bg=framebg, fg=framefg, command=selection)
r1.place(x=150, y=150)

r2 = Radiobutton(obj, text="Female", variable=radio, value=2, bg=framebg, fg=framefg, command=selection)
r2.place(x=200, y=150)

religion = StringVar()
religion_entry = Entry(obj, textvariable=religion, width=20, font="arial 10")
religion_entry.place(x=630, y=100)

skill = StringVar()
skill_entry = Entry(obj, textvariable=skill, width=20, font="arial 10")
skill_entry.place(x=630, y=150)

Class = Combobox(obj, values=['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12'], font='Roboto 10', width=17, state='r')
Class.place(x=630, y=50)
Class.set("Select Class")

#parent's details
obj2 = LabelFrame(root, text="Parent's Details", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=250, relief=GROOVE)
obj2.place(x=30, y=470)

Label(obj2, text="Father's Name:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=50)
Label(obj2, text="Occupation:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=100)

F_Name = StringVar()
f_entry = Entry(obj2, textvariable=F_Name, width=20, font="arial 10")
f_entry.place(x=160, y=50)

father_occupation = StringVar()
fo_entry = Entry(obj2, textvariable=father_occupation, width=20, font="arial 10")
fo_entry.place(x=160, y=100)

Label(obj2, text="Mother's Name:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=50)
Label(obj2, text="Occupation:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=100)


m_Name = StringVar()
m_entry = Entry(obj2, textvariable=m_Name, width=20, font="arial 10")
m_entry.place(x=630, y=50)

mother_occupation = StringVar()
mo_entry = Entry(obj2, textvariable=mother_occupation, width=20, font="aarial 10")
mo_entry.place(x=630, y=100)

#image
f = Frame(root, bd=4, bg="black", width=200, height=200, relief=GROOVE)
f.place(x=1000, y=150)

img = PhotoImage(file="C:/Users/colis/Documents/tetris/registration/images/profile4.png")
lbl = Label(f, bg="black", image=img)
lbl.place(x=4, y=4)


#button
Button(root, text="Upload", width=19, height=2, font="arial 12 bold", bg="lightblue", command=showimage).place(x=1000, y=370)

savebutton = Button(root, text="Save", width=19, height=2, font="arial 12 bold", bg="lightgreen", command=Save)
savebutton.place(x=1000, y=450)

Button(root, text="Reset", width=19, height=2, font="arial 12 bold", bg="lightpink", command=clear).place(x=1000, y=530)

Button(root, text="Exit", width=19, height=2, font="arial 12 bold", bg="grey", command=Exit).place(x=1000, y=610)


root.mainloop()